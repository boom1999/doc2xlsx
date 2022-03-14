# -*- coding: utf-8 -*-
# @Time : 2022/3/4 14:43
# @Author : lingz
# @Software: PyCharm

import copy
from docx import Document
import xlrd
import pandas as pd
from pandas import DataFrame
import os
import re
import shutil
from openpyxl import load_workbook
import difflib


def get_docx_files_path(input_path):
    """
    input dirs of the docx_s, then listdir input_path-->dirs_path-->docx_files_path.
    :param input_path: the path containing the whole input files.
    :return: the path of the dirs and the docx_file which need to tidy up.
    """
    docx_files_path = []
    docx_files = []
    dirs_path = os.listdir(input_path)
    for index in range(0, len(dirs_path)):
        dirs_path[index] = str(os.path.join(input_path, dirs_path[index]))
        files = os.listdir(dirs_path[index])
        for file in files:
            if os.path.splitext(file)[1] == ".docx":
                docx_files.append(file)
        docx_files_path.append(os.path.join(dirs_path[index] + '/', docx_files[index]))
    return dirs_path, docx_files_path


def read_docx(read_path):
    """

    :param read_path: .docx files.
    :return: list_ -- contain Dicts = {":", ";"...}.
    """
    input_path = copy.deepcopy(read_path)
    docx = Document(input_path)
    table_s = docx.tables

    # 从文件路径拆分出文件名，从文件名中按照‘-’拆出编码 code_
    (filepath, filename) = os.path.split(input_path)
    code_list = filename.split('-')[0:3]
    code_ = code_list[0] + '-' + code_list[1] + '-' + code_list[2]

    list_ = []
    for table in table_s:
        for j in range(1, len(table.rows)):
            dict_ = {'序号': table.cell(j, 0).text,
                     '展品名称': table.cell(j, 1).text,
                     '所属领域': table.cell(j, 2).text,
                     '展品形式': table.cell(j, 3).text,
                     '展品单位': table.cell(j, 4).text,
                     '联系人': table.cell(j, 5).text,
                     '联系电话': table.cell(j, 6).text,
                     '编码': code_}
            if dict_['展品名称'] != '':
                list_.append(dict_)
                print("Append successfully: row " + str(j))
    return list_


def add_type_region(list_in, map_tables_path):
    """

    :param map_tables_path: path of the map, original list and their types and regions.
    :param list_in: putin original list which need to add type and region.
    :return: final list -- containing type and region.
    """

    # Use copy.deepcopy to avoid to change list_in outsider
    list_out = copy.deepcopy(list_in)
    # 这里需要拆分目录读取编码
    code_ = '编码'
    flat_ = '展品单位'
    # flat_simple = '单位简称'
    type_ = '类型'
    origin_ = '地区'
    # Attention! 'map_tables_path' can't have head index, pd.read_excel can't read.
    # Or set 'header= '
    map_data = pd.read_excel(map_tables_path, index_col='编码', header=1, sheet_name='Sheet 1')
    code_list = list(map_data.index)
    for j in range(0, len(list_out)):
        if list_out[j][code_] not in code_list:
            print(list_out[j][flat_] + "出现编码错误，请手动修改!")
            continue
        list_out[j][flat_] = map_data.loc[list_out[j][code_], flat_]
        # list_out[j][flat_simple] = map_data.loc[list_out[j][code_], flat_simple]
        list_out[j][type_] = map_data.loc[list_out[j][code_], type_]
        list_out[j][origin_] = map_data.loc[list_out[j][code_], origin_]

    return list_out


def export2excel(export, out):
    """

    :param export: list_sum -- contain Dicts = {":", ";"...} +  {":", ";"...} + ...
    :param out: outfile .xlsx file.
    :return:
    """
    pf = pd.DataFrame(list(export))

    # Redefine column labels
    order = ['编码', '序号', '展品名称', '所属领域', '展品形式', '展品单位', '类型', '地区', '联系人', '联系电话']
    pf = pf[order]

    file_path = pd.ExcelWriter(out)
    pf.fillna(' ', inplace=True)
    pf.to_excel(file_path, encoding='utf-8', index=False)
    file_path.save()


def mkdir_classify(classify_path):
    """

    :param classify_path: The path need to be created.
    :return: The classify_path.
    """
    classify_path = classify_path.strip()
    classify_path = classify_path.rstrip("\\")

    isExists = os.path.exists(classify_path)
    if not isExists:
        os.makedirs(classify_path)
        print("分类路径'" + str(classify_path) + "'创建成功！")
        return classify_path
    else:
        print("分类路径'" + str(classify_path) + "'已存在，跳过创建！")
        return classify_path


def classify(summary_path, dirs_path, classify_path):
    """

    :param summary_path: The original statistical table path.
    :param dirs_path: The file summary path.
    :param classify_path: The output path.
    :return:
    """
    path = copy.deepcopy(summary_path)
    data = pd.read_excel(path, index_col='所属领域')
    summary_type_list = list(set(data.index))
    for summary_type in summary_type_list:
        type_path = mkdir_classify(os.path.join(classify_path + '/', str(summary_type)))
        # 注意：这里data.loc[[summary_type], :])第一个参数必须要加[]，否则会解析成series.Series，需要使用frame.DataFrame保存
        target_path = str(type_path) + '/'
        temp_target_path = copy.deepcopy(target_path)
        data.loc[[summary_type], :].to_excel(target_path + summary_type + '汇总表.xlsx')
        print(summary_type + '汇总表.xlsx 创建成功')
        print("--------------------------------------------------------------------")
        print("Moving file structure......")
        temp_data = data.loc[[summary_type], :]
        for i in range(0, temp_data.shape[0]):
            code_ = temp_data.iloc[i, 0]
            num_ = temp_data.iloc[i, 1]
            name_ = temp_data.iloc[i, 4]
            for dir_path in dirs_path:
                f_path, f_name = os.path.split(dir_path)
                if re.match(code_ + '.+', f_name) is not None:
                    temp_dirs_path = os.listdir(dir_path)
                    for temp_dir_path in temp_dirs_path:
                        # 用正则从路径中拆出数字, like: 10、源牌电磁水表（实物） to ['10', '']
                        path_list = re.split("[^0-9]+", temp_dir_path)
                        for path in path_list:
                            if path == str(num_) and os.path.splitext(temp_dir_path)[1] != ".docx" \
                                    and os.path.splitext(temp_dir_path)[1] != ".PDF":
                                src_path = os.path.join(dir_path + '/', temp_dir_path)
                                target_path = os.path.join(target_path, '(' + str(name_) + ')' + temp_dir_path)
                                if not os.path.exists(target_path):
                                    shutil.copytree(src_path, target_path)
                                    target_path = copy.deepcopy(temp_target_path)
                                    print(str(src_path) + '-->' + str(target_path) + "移动成功!")
                                else:
                                    target_path = copy.deepcopy(temp_target_path)
                                    print("----------------目标路径%s已存在，跳过！------------------" % target_path)
                                break
                    break
        print("--------------------------------------------------------------------")


def cell_handling(summary_path):
    """
    Line breaks and '\' existing in cells in the original table will affect path creation and need to be replaced.
    :param summary_path: file path to process.
    :return:
    """
    wb = load_workbook(summary_path)
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            old = ws.cell(i, j).value
            if old is not None:
                ws.cell(i, j).value = old.strip().replace(' ', '').replace("\n", "").replace("/", "和")
    wb.save(summary_path)
    wb.close()
    print("换行符和反斜杠处理完成!")


def string_similar(s1, s2):
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()


def classify_2(src_path, dst_path):
    """

    :param src_path:
    :param dst_path:
    :return:
    """
    temp_dirs_path = os.listdir(src_path)
    temp_src_path = copy.deepcopy(src_path)
    temp_dst_path = copy.deepcopy(dst_path)

    for temp_dir_path in temp_dirs_path:
        temp_src_path = os.path.join(temp_src_path + '/', temp_dir_path)
        temp_dst_path = os.path.join(temp_dst_path + '/', temp_dir_path)
        if not os.path.exists(temp_dst_path):
            if os.path.splitext(temp_src_path)[1] != ".xlsx":
                shutil.copytree(temp_src_path, temp_dst_path)
            else:
                shutil.copyfile(temp_src_path, temp_dst_path)
            print(str(temp_src_path) + '-->' + str(temp_dst_path) + "移动成功!")
            temp_src_path = copy.deepcopy(src_path)
            temp_dst_path = copy.deepcopy(dst_path)
        else:
            print("----------------目标路径%s已存在，跳过！------------------" % temp_dst_path)
            temp_src_path = copy.deepcopy(src_path)
            temp_dst_path = copy.deepcopy(dst_path)


def input_index():
    """

    :return: type_list: [num1, num2, ...]
    """
    input_data = input()
    input_data = input_data.strip()
    input_list = input_data.split(" ")
    input_list = list(map(int, input_list))
    for data in input_list:
        if data < 1 or data > 7:
            print("越界，请重新输入！")
            input_list = input_index()
    return input_list


def merge_type_xlsx(static_path):
    """
    Merge sub-domain summary tables.
    :return:
    """
    static_type_path_list = os.listdir(static_path)
    for index in range(0, len(static_type_path_list)):
        file_name = static_type_path_list[index]
        static_type_path_list[index] = os.path.join(static_path + '/', static_type_path_list[index])
        DFs = []
        dirs_list = os.listdir(static_type_path_list[index])
        out_summary_path = os.path.join(static_type_path_list[index] + '/', file_name + '.xlsx')
        if os.path.exists(out_summary_path):
            os.remove(out_summary_path)
        for dir_path in dirs_list:
            if os.path.splitext(dir_path)[1] == ".xlsx":
                file_path = os.path.join(static_type_path_list[index] + '/', dir_path)
                df = pd.read_excel(file_path)
                DFs.append(df)
                os.remove(file_path)
        writer = pd.ExcelWriter(out_summary_path)
        pd.concat(DFs).to_excel(writer, index=False)
        writer.save()
        print(file_name+"合并完毕")


def merge_summary_xlsx(output_path, deleting=True):
    """

    :param output_path: Original xlsx output path.
    :param deleting: Whether to delete the original xlsx.
    :return:
    """
    xlsx_name_list = os.listdir(output_path)
    xlsx_path_list = []
    DFs = []
    out_summary_path = os.path.join(output_path + '/', 'summary.xlsx')
    for index in range(0, len(xlsx_name_list)):
        xlsx_path_list.append(os.path.join(output_path + '/', xlsx_name_list[index]))
    for xlsx_path in xlsx_path_list:
        if os.path.splitext(xlsx_path)[1] == ".xlsx":
            if xlsx_path == out_summary_path:
                os.remove(out_summary_path)
                continue
            df = pd.read_excel(xlsx_path)
            DFs.append(df)
            if deleting:
                os.remove(xlsx_path)
    writer = pd.ExcelWriter(out_summary_path)
    pd.concat(DFs).to_excel(writer, index=False)
    writer.save()
    print("合并完毕！\n")
